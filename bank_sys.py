import openpyxl as xl
file = xl.load_workbook('bank_accounts.xlsx')
sheet = file.active

class bank_account:
    def __init__(self, name, acc_no, ifsc_code, bank_balance, row_num):
        self.name = f"name of the bank holder: {name}\n"
        self.acc_no = f"Account number of bank holder: {acc_no}\n"
        self.ifsc_code = f"IFSC code of bank: {ifsc_code}\n"
        self.bank_balance = float(bank_balance)
        self.row = row_num

    def dep(self, deposit):
        self.bank_balance += deposit
        sheet.cell(row=self.row, column=4).value = self.bank_balance
        file.save('bank_accounts.xlsx')

    def wd(self, withdraw):
        self.bank_balance -= withdraw
        sheet.cell(row=self.row, column=4).value = self.bank_balance
        file.save('bank_accounts.xlsx')
        if self.bank_balance < 0:
            print("Insufficient balance")
            self.bank_balance += withdraw
            sheet.cell(row=self.row, column=4).value = self.bank_balance

    def cb(self):
        print(self.bank_balance)

    def display(self):
        print(self.name, self.acc_no, self.ifsc_code, self.bank_balance)

max_row = sheet.max_row

def fetch_bank_details(argument):
    match argument:
        case 0:
            return "Enter the name of the bank holder: "
        case 1:
            return "Enter the account number of the bank holder: "

n = int(input("Enter 0 for name or 1 for account number: "))
prompt = fetch_bank_details(n)
user_input = input(prompt)

found = False
for row in range(1, max_row + 1):
    cell = sheet.cell(row=row, column=1 if n == 0 else 2)
    if cell.value == user_input:
        name = sheet.cell(row=row, column=1).value
        acc = sheet.cell(row=row, column=2).value
        ifsc = sheet.cell(row=row, column=3).value
        bl = sheet.cell(row=row, column=4).value
        obj1 = bank_account(name, acc, ifsc, bl, row)
        obj1.display()
        num = input("Enter 1 for deposit, 2 for withdraw: ")
        if num == '1':
            deposit = float(input("Enter the amount to deposit: "))
            obj1.dep(deposit)
            print("Deposit successful.")
        elif num == '2':
            withdraw = float(input("Enter the amount to withdraw: "))
            obj1.wd(withdraw)
            print("Withdrawal successful.")
        else:
            print("Invalid option selected.")
        check = input("Do you want to check your bank balance? (yes/no): ")
        if check.lower() == 'yes':
            obj1.cb()
        else:
            print("Thank you for using our service.")
        found = True
        break

if not found:
    print("Bank account not found.")
